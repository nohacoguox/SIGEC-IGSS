"""Regenera ficha con reglas Palín–DAF. Escribe en ./punto3_EJEMPLO_COMPLETO y opcionalmente copia a Documents."""
from __future__ import annotations

import csv
import shutil
from datetime import date, timedelta
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False

OUT = Path(__file__).resolve().parent / "punto3_EJEMPLO_COMPLETO"
DOCS = Path(r"C:\Users\estua\OneDrive\Documents\UMG\10MO\pg2\PRIMERA ENTREGA\punto3_EJEMPLO_COMPLETO")


def networkdays(start: date, end: date) -> int:
    if end < start:
        return 0
    n = 0
    d = start
    while d <= end:
        if d.weekday() < 5:
            n += 1
        d += timedelta(days=1)
    return n


def add_business_days(start: date, days: int) -> date:
    if days <= 0:
        return start
    d = start
    left = days
    while left > 0:
        d += timedelta(days=1)
        if d.weekday() < 5:
            left -= 1
    return d


def clamp_send_day(year: int, month: int, preferred: int) -> date:
    day = max(16, min(22, preferred))
    d = date(year, month, day)
    while d.weekday() >= 5:
        if d.day < 22:
            d += timedelta(days=1)
        else:
            d -= timedelta(days=1)
    if d.day < 16:
        d = date(year, month, 16)
        while d.weekday() >= 5:
            d += timedelta(days=1)
    if d.day > 22:
        d = date(year, month, 22)
        while d.weekday() >= 5:
            d -= timedelta(days=1)
    return d


def next_month_send(year: int, month: int, preferred: int = 18) -> date:
    if month == 12:
        return clamp_send_day(year + 1, 1, preferred)
    return clamp_send_day(year, month + 1, preferred)


RAW = [
    ("LB-01", "SIAF-03-2025", "EXP-01-2025", "Insumos", 2025, 1, 17, 9, 2, "Observacion", 1, 0, 1, False,
     "Observacion menor; Palin corrigio en el mismo ciclo mensual"),
    ("LB-02", "SIAF-05-2025", "EXP-02-2025", "Servicio", 2025, 1, 20, 11, 3, "Observacion", 3, 0, 1, False,
     "Servicio: justificacion del SIAF no coincidia con criterio DAF; se corrigio redaccion"),
    ("LB-03", "SIAF-07-2025", "EXP-03-2025", "Activo fijo", 2025, 1, 22, 14, 3, "Aprobacion", 0, 0, 0, False,
     "Aprobado en primer analisis del expediente fisico"),
    ("LB-04", "SIAF-09-2025", "EXP-04-2025", "Servicio", 2025, 2, 18, 10, 3, "Observacion", 4, 1, 2, True,
     "Servicio: multiples correcciones de descripcion; no se subsano a tiempo y quedo para marzo"),
    ("LB-05", "SIAF-11-2025", "EXP-05-2025", "Insumos", 2025, 2, 19, 8, 2, "Aprobacion", 0, 0, 0, False,
     "Sin observaciones en primer analisis"),
    ("LB-06", "SIAF-14-2025", "EXP-06-2025", "Insumos", 2025, 2, 21, 12, 2, "Observacion", 2, 0, 1, False,
     "Faltaba anexo; corregido en el mismo mes"),
    ("LB-07", "SIAF-16-2025", "EXP-07-2025", "Servicio", 2025, 3, 17, 9, 3, "Observacion", 3, 0, 1, False,
     "Servicio: forma de justificar el gasto objetada; se reescribio SIAF"),
    ("LB-08", "SIAF-18-2025", "EXP-08-2025", "Activo fijo", 2025, 3, 20, 13, 3, "Observacion", 2, 0, 1, False,
     "Especificacion tecnica incompleta; corregida antes del cierre mensual"),
    ("LB-09", "SIAF-21-2025", "EXP-09-2025", "Servicio", 2025, 3, 21, 11, 3, "Rechazo", 4, 1, 2, True,
     "Servicio: correcciones no aplicadas a tiempo en Palin; rechazo y reingreso en abril"),
    ("LB-10", "SIAF-23-2025", "EXP-10-2025", "Insumos", 2025, 4, 16, 7, 2, "Aprobacion", 0, 0, 0, False,
     "Aprobado sin devoluciones"),
    ("LB-11", "SIAF-25-2025", "EXP-11-2025", "Servicio", 2025, 4, 18, 10, 3, "Observacion", 3, 0, 1, False,
     "Servicio: palabras/descripcion del SIAF ajustadas segun criterio DAF"),
    ("LB-12", "SIAF-28-2025", "EXP-12-2025", "Insumos", 2025, 4, 22, 12, 2, "Observacion", 1, 0, 1, False,
     "Observacion por cotizacion; subsanada en el mes"),
    ("LB-13", "SIAF-30-2025", "EXP-13-2025", "Servicio", 2025, 5, 19, 11, 3, "Rechazo", 5, 1, 2, True,
     "Servicio: mala produccion por rechazo; pendiente para junio por no corregir a tiempo"),
    ("LB-14", "SIAF-32-2025", "EXP-14-2025", "Activo fijo", 2025, 5, 20, 9, 2, "Aprobacion", 0, 0, 0, False,
     "Aprobado en primer analisis"),
    ("LB-15", "SIAF-35-2025", "EXP-15-2025", "Insumos", 2025, 5, 21, 8, 3, "Observacion", 2, 0, 1, False,
     "Correccion documental menor en el mismo ciclo"),
    ("LB-16", "SIAF-37-2025", "EXP-16-2025", "Servicio", 2025, 6, 17, 10, 3, "Observacion", 3, 0, 1, False,
     "Servicio: justificacion reelaborada para cumplir criterio de redaccion DAF"),
    ("LB-17", "SIAF-40-2025", "EXP-17-2025", "Insumos", 2025, 6, 18, 6, 2, "Aprobacion", 0, 0, 0, False,
     "Sin observaciones"),
    ("LB-18", "SIAF-42-2025", "EXP-18-2025", "Servicio", 2025, 6, 20, 12, 3, "Observacion", 4, 0, 2, True,
     "Servicio: varias vueltas de redaccion; no cerro en junio y paso a julio"),
]


def build_rows():
    rows = []
    for (
        idc, siaf, exp, tipo, year, month, day_envio, inicio_offset, dias_analisis,
        tipo1, obs, rech, ciclos, pasa_mes, nota,
    ) in RAW:
        puesta = clamp_send_day(year, month, day_envio)
        inicio = puesta - timedelta(days=inicio_offset)
        while inicio.weekday() >= 5:
            inicio -= timedelta(days=1)
        primera = add_business_days(puesta, dias_analisis)

        if tipo1 == "Aprobacion":
            aprob = primera
            devuelto = "No"
        elif pasa_mes:
            reenvio = next_month_send(year, month, 21 if ciclos >= 2 else 18)
            aprob = add_business_days(reenvio, dias_analisis)
            devuelto = "Sí"
        else:
            dias_correccion = 4 if tipo == "Servicio" else 3
            aprob = add_business_days(primera, dias_correccion + max(0, ciclos - 1) * 2)
            devuelto = "Sí"

        dias_ciclo = networkdays(inicio, aprob)
        dias_resp = networkdays(puesta, primera)

        t1 = t2 = t3 = t5 = "Sí"
        t4 = "Sí" if ciclos == 0 else "No"
        if pasa_mes:
            t4 = "No"
        t6 = "No"
        t7 = "No" if ciclos >= 1 or pasa_mes else "Sí"
        cumple = "Sí" if all(x == "Sí" for x in (t1, t2, t3, t4, t5)) else "No"
        minutos = 3 + (0 if ciclos == 0 else 5) + (3 if tipo == "Servicio" else 0) + (4 if pasa_mes else 0) + obs

        rows.append({
            "id": idc, "siaf": siaf, "exp": exp, "tipo": tipo,
            "inicio": inicio, "puesta": puesta, "primera": primera, "tipo1": tipo1, "aprob": aprob,
            "dias_ciclo": dias_ciclo, "dias_resp": dias_resp, "obs": obs, "rech": rech,
            "devuelto": devuelto, "ciclos": ciclos,
            "fuente": "Expediente fisico + planilla Compras Palin (envio mensual 16-22)",
            "nota": nota, "t1": t1, "t2": t2, "t3": t3, "t4": t4, "t5": t5, "t6": t6, "t7": t7,
            "cumple": cumple, "minutos": minutos, "pasa_mes": pasa_mes,
        })
    return rows


def avg(vals):
    return sum(vals) / len(vals) if vals else 0


def write_outputs(out: Path, rows, ind):
    out.mkdir(parents=True, exist_ok=True)
    headers = [
        "ID_caso", "Numero_SIAF", "Numero_Expediente", "Tipo_tramite", "Fecha_inicio",
        "Fecha_puesta_revision_DAF", "Fecha_primera_resolucion_DAF", "Tipo_primera_resolucion",
        "Fecha_aprobacion_DAF", "Dias_habiles_ciclo_total", "Dias_habiles_primera_respuesta_DAF",
        "N_observaciones", "N_rechazos_formales", "Devuelto_al_menos_1_vez",
        "N_ciclos_revision_correccion_reenvio", "Fuente_datos", "Observaciones_registro",
    ]
    headers2 = [
        "ID_caso", "T1_Identificacion_clara", "T2_Fecha_inicio", "T3_Responsable_area",
        "T4_Historial_devoluciones", "T5_Estado_final_DAF", "T6_Versiones_distinguibles",
        "T7_Fecha_cambios", "Cumple_global_T1_a_T5", "Minutos_localizar_ultima_version_estado",
        "Observaciones",
    ]

    with (out / "03_Ficha_LLENA.csv").open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(headers)
        for r in rows:
            w.writerow([
                r["id"], r["siaf"], r["exp"], r["tipo"],
                r["inicio"].strftime("%d/%m/%Y"), r["puesta"].strftime("%d/%m/%Y"),
                r["primera"].strftime("%d/%m/%Y"), r["tipo1"], r["aprob"].strftime("%d/%m/%Y"),
                r["dias_ciclo"], r["dias_resp"], r["obs"], r["rech"], r["devuelto"], r["ciclos"],
                r["fuente"], r["nota"],
            ])

    with (out / "04_Cotejo_LLENO.csv").open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(headers2)
        for r in rows:
            w.writerow([
                r["id"], r["t1"], r["t2"], r["t3"], r["t4"], r["t5"], r["t6"], r["t7"], r["cumple"],
                r["minutos"],
                "Envio fisico 16-22; DAF analiza 2-3 dias; servicios con mas correcciones de redaccion/justificacion",
            ])

    if HAS_XLSX:
        wb = Workbook()
        fill = PatternFill("solid", fgColor="1F4E79")
        font_h = Font(color="FFFFFF", bold=True)
        thin = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        ws = wb.active
        ws.title = "Ficha_LineaBase"
        for c, h in enumerate(headers, 1):
            cell = ws.cell(1, c, h)
            cell.fill = fill
            cell.font = font_h
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        for i, r in enumerate(rows, 2):
            vals = [r["id"], r["siaf"], r["exp"], r["tipo"], r["inicio"], r["puesta"], r["primera"], r["tipo1"], r["aprob"],
                    r["dias_ciclo"], r["dias_resp"], r["obs"], r["rech"], r["devuelto"], r["ciclos"], r["fuente"], r["nota"]]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(i, c, v)
                cell.border = thin
                if isinstance(v, date):
                    cell.number_format = "DD/MM/YYYY"
        ws2 = wb.create_sheet("Lista_Cotejo")
        for c, h in enumerate(headers2, 1):
            cell = ws2.cell(1, c, h)
            cell.fill = fill
            cell.font = font_h
        for i, r in enumerate(rows, 2):
            for c, v in enumerate([r["id"], r["t1"], r["t2"], r["t3"], r["t4"], r["t5"], r["t6"], r["t7"], r["cumple"], r["minutos"],
                                   "Envio 16-22; analisis DAF 2-3 dias; servicios con mas correcciones"], 1):
                ws2.cell(i, c, v).border = thin
        ws3 = wb.create_sheet("Indicadores_resueltos")
        ws3["A1"] = "Indicador"
        ws3["B1"] = "Valor"
        for i, (a, b) in enumerate([
            ("n", ind["n"]),
            ("Periodo", ind["periodo"]),
            ("Tiempo medio ciclo total", round(ind["ciclo"], 2)),
            ("Tiempo medio 1a respuesta DAF", round(ind["resp"], 2)),
            ("Promedio ciclos", round(ind["ciclos"], 2)),
            ("% devueltos >=1", round(ind["pct_dev"], 2)),
            ("Promedio observaciones", round(ind["obs"], 2)),
            ("Rechazos / 100", round(ind["rech100"], 2)),
            ("% trazabilidad", round(ind["pct_traz"], 2)),
            ("Minutos busqueda", round(ind["mins"], 2)),
            ("Casos que pasaron al mes siguiente", ind["pasan_mes"]),
            ("% devoluciones en Servicio", round(ind["pct_serv_dev"], 2)),
            ("Obs promedio Servicio", round(ind["obs_serv"], 2)),
            ("Obs promedio otros", round(ind["obs_otros"], 2)),
            ("AVISO", "DATOS ILUSTRATIVOS - reemplazar con reales"),
        ], 2):
            ws3.cell(i, 1, a)
            ws3.cell(i, 2, b)
        ws3.column_dimensions["A"].width = 45
        ws3.column_dimensions["B"].width = 40
        for wsx in (ws, ws2):
            wsx.freeze_panes = "B2"
            wsx.row_dimensions[1].height = 40
            for col in wsx.columns:
                wsx.column_dimensions[col[0].column_letter].width = 16
        wb.save(out / "03_Instrumento_LineaBase_LLENO.xlsx")

    (out / "00_LEE_PRIMERO.md").write_text(
        f"""# Simulacion actualizada

Reglas: envio 16-22; DAF 2-3 dias; sin correccion a tiempo -> mes siguiente; Servicios con mas correcciones.

- n={ind['n']}
- ciclo={ind['ciclo']:.2f}
- 1a respuesta={ind['resp']:.2f}
- ciclos={ind['ciclos']:.2f}
- %devueltos={ind['pct_dev']:.2f}%
- obs={ind['obs']:.2f}
- rech/100={ind['rech100']:.2f}
- trazabilidad={ind['pct_traz']:.2f}%
- mins={ind['mins']:.2f}
- pasan_mes={ind['pasan_mes']}
- obs_serv={ind['obs_serv']:.2f}
- obs_otros={ind['obs_otros']:.2f}
""",
        encoding="utf-8",
    )

    (out / "02_TEXTO_TESIS_CON_NUMEROS.md").write_text(
        f"""# Resultados simulacion (ilustrativos)

| Indicador | Valor |
|-----------|------:|
| n | {ind['n']} |
| Periodo | {ind['periodo']} |
| Tiempo medio ciclo total | {ind['ciclo']:.2f} |
| 1a respuesta DAF | {ind['resp']:.2f} |
| Promedio ciclos | {ind['ciclos']:.2f} |
| % devueltos >=1 | {ind['pct_dev']:.2f}% |
| Promedio observaciones | {ind['obs']:.2f} |
| Rechazos/100 | {ind['rech100']:.2f} |
| % trazabilidad | {ind['pct_traz']:.2f}% |
| Minutos busqueda | {ind['mins']:.2f} |
| Pasaron al mes siguiente | {ind['pasan_mes']} |
| Obs promedio Servicios | {ind['obs_serv']:.2f} |
| Obs promedio otros | {ind['obs_otros']:.2f} |
""",
        encoding="utf-8",
    )


def main():
    rows = build_rows()
    for r in rows:
        assert 16 <= r["puesta"].day <= 22, r
        assert 2 <= r["dias_resp"] <= 3, (r["id"], r["dias_resp"])
    n = len(rows)
    ind = {
        "n": n,
        "periodo": f"{min(r['inicio'] for r in rows).strftime('%d/%m/%Y')} – {max(r['aprob'] for r in rows).strftime('%d/%m/%Y')}",
        "ciclo": avg([r["dias_ciclo"] for r in rows]),
        "resp": avg([r["dias_resp"] for r in rows]),
        "ciclos": avg([r["ciclos"] for r in rows]),
        "pct_dev": sum(1 for r in rows if r["devuelto"] == "Sí") / n * 100,
        "obs": avg([r["obs"] for r in rows]),
        "rech100": sum(r["rech"] for r in rows) / n * 100,
        "pct_traz": sum(1 for r in rows if r["cumple"] == "Sí") / n * 100,
        "mins": avg([r["minutos"] for r in rows]),
        "pct_serv_dev": sum(1 for r in rows if r["tipo"] == "Servicio" and r["devuelto"] == "Sí")
        / max(1, sum(1 for r in rows if r["tipo"] == "Servicio")) * 100,
        "obs_serv": avg([r["obs"] for r in rows if r["tipo"] == "Servicio"]),
        "obs_otros": avg([r["obs"] for r in rows if r["tipo"] != "Servicio"]),
        "pasan_mes": sum(1 for r in rows if r["pasa_mes"]),
    }
    write_outputs(OUT, rows, ind)
    try:
        DOCS.mkdir(parents=True, exist_ok=True)
        for p in OUT.iterdir():
            shutil.copy2(p, DOCS / p.name)
        print("COPIADO_A", DOCS)
    except Exception as e:
        print("COPY_WARN", e)
    print("OUT", OUT)
    for k, v in ind.items():
        print(f"{k}={v}")


if __name__ == "__main__":
    main()
